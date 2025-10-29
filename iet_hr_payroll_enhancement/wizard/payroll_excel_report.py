from odoo import models, fields
import base64
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


class PayrollExcelReportWizard(models.TransientModel):
    _name = 'payroll.excel.report.wizard'
    _description = 'Excel Payroll Report Wizard'

    batch_id = fields.Many2one('hr.payslip.run', string='Batch', required=True)

    def print_report(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Payroll Report"

        header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        headers = [
            ("Employee #", "الرقم الوظيفى"),
            ("Employee Name", "إسم الموظف"),
            ("Account #", "رقم الحساب"),
            ("Bank", "البنك"),
            ("Payment Method", "طريقة الصرف"),
            ("Gross Salary", "إجمالي الراتب"),
            ("Net Amount", "المبلغ الصافي"),
            ("Legal #", "رقم الهوية/الاقامة"),
            ("Basic Wage", "الراتب الاساسى"),
            ("Housing Allowance", "بدل السكن"),
            ("Transportation Allowance", "بدل النقل"),
            ("Food Allowance", "بدل الطعام"),
            ("Phone Allowance", "بدل الهاتف"),
            ("Natural Of Work Allowance", "بدل طبيعة العمل"),
            ("Other Earnings", "دخل اخر"),
            ("Social Insurance Deduction", "خصم التأمينات الاجتماعية"),
            ("Other Deductions", "خصومات أخرى"),
            ("Joining Date", "تاريخ الدخول"),
        ]

        column_widths = [25, 27, 25, 25, 27, 25, 25, 27, 27, 27, 34,34,34,40, 27, 36, 27, 27]
        for i, width in enumerate(column_widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = width

        for idx, (eng, ar) in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=idx)
            cell.value = f"{eng}\n{ar}"
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        sheet.row_dimensions[1].height = 40
        employees = []
        for slip in self.batch_id.slip_ids:
            employee = slip.employee_id
            if employee.id not in employees:
                employees.append(employee.id)
                payslip_ids = self.batch_id.slip_ids.filtered(lambda x: x.employee_id.id == employee.id)
                data = [
                    employee.employee_number,
                    employee.name,
                    employee.bank_iban_number,
                    employee.bank_name,
                    ', '.join(employee.category_ids.mapped('name')),
                    # slip.contract_id.wage,
                    sum(payslip_ids.mapped('line_ids').filtered(lambda line: line.code.upper() == 'GROSS').mapped('amount')),
                    sum(payslip_ids.mapped('line_ids').filtered(lambda line: line.code.lower() == 'net').mapped(
                        'amount')),
                    employee.identification_id,
                    sum(payslip_ids.mapped('line_ids').filtered(lambda line: line.code.lower() == 'basic').mapped(
                        'amount')),
                    sum(payslip_ids.mapped('line_ids').filtered(lambda line: line.name.lower() == 'housing allowance').mapped(
                        'amount')),
                    sum(payslip_ids.mapped('line_ids').filtered(
                        lambda line: line.name.lower() == 'transportation allowance').mapped('amount')),
                    sum(payslip_ids.mapped('line_ids').filtered(
                        lambda line: line.code.lower() == 'foall').mapped('amount')),
                    sum(payslip_ids.mapped('line_ids').filtered(
                        lambda line: line.name.lower() == 'phone allowance').mapped('amount')),
                    sum(payslip_ids.mapped('line_ids').filtered(
                        lambda line: line.name.lower() == 'natural of work allowance').mapped('amount')),
                    sum(payslip_ids.mapped('line_ids').filtered(lambda line: line.name.lower() == 'other allowances').mapped(
                        'amount')),
                    sum(payslip_ids.mapped('line_ids').filtered(lambda line: line.code == 'GOSI_EMP').mapped('amount')),
                    sum(payslip_ids.mapped('line_ids').filtered(
                        lambda line: line.category_id.name.lower() == 'deduction' and line.code != 'GOSI_EMP').mapped(
                        'amount')),
                    employee.first_contract_date,
                ]
                sheet.append(data)

        stream = io.BytesIO()
        workbook.save(stream)
        file_data = base64.b64encode(stream.getvalue())
        stream.close()

        attachment = self.env['ir.attachment'].create({
            'name': f'Payroll_Report_{self.batch_id.name}.xlsx',
            'type': 'binary',
            'datas': file_data,
            'res_model': 'hr.payslip.run',
            'res_id': self.batch_id.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }

    def cancel(self):
        return {'type': 'ir.actions.act_window_close'}
